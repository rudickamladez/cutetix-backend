"""Module for easier event management"""
from app.models import Event, TicketStatusEnum
from app.schemas import extra
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from tempfile import NamedTemporaryFile
from io import BytesIO
import sys
import re


def get_event_capacity_summary(event: Event):
    if not event:
        print(
            "There is not event",
            file=sys.stderr,
        )

    # Prepare response
    cs = extra.CapacitySummary()

    for tg in event.ticket_groups:
        cs.total += tg.capacity
        tg_free = tg.capacity
        for t in tg.tickets:
            # Increase per status
            if t.status == TicketStatusEnum.new:
                cs.reserved += 1
            # elif t.status == TicketStatusEnum.confirmed:
            elif t.status == TicketStatusEnum.paid:
                cs.paid += 1
            elif t.status == TicketStatusEnum.cancelled:
                cs.cancelled += 1

            # Decrase free tickets
            if not t.status == TicketStatusEnum.cancelled:
                tg_free -= 1
        cs.free += tg_free
    return cs


def get_event_xlsx(event: Event):

    # Create workbook
    wb = Workbook(iso_dates=True)

    for tg in event.ticket_groups:
        # Create worksheet for ticket group
        title = re.sub(INVALID_TITLE_REGEX, '_', tg.name)
        ws = wb.create_sheet(title=title)

        # Create heading
        ws.append([
            "Lastname",
            "Firstname",
            "E-mail",
            "Description",
            "Status",
            "Ticket ID",
            "Order date",
        ])

        for t in tg.tickets:
            # Add row with ticket
            ws.append([
                t.lastname,
                t.firstname,
                t.email,
                t.description,
                str(t.status),
                t.id,
                t.order_date,
            ])

    # Remove default blank sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Return workbook
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        return BytesIO(tmp.read())


def get_event_xlsx_for_libor(event: Event):

    # Load formatted workbook
    wb = load_workbook(
        filename="./app/services/cutetix-event-formatted-for-Libor.xlsx")

    # Sort time groups
    tgs = sorted(event.ticket_groups, key=lambda tg: tg.name.lower())

    FIRST_ROW_TO_UPDATE = 3
    for tg in tgs:
        # Create worksheet for ticket group
        title = re.sub(INVALID_TITLE_REGEX, '_', tg.name)
        default_ws = wb["default"]
        ws = wb.copy_worksheet(from_worksheet=default_ws)
        ws.title = title

        # Update title of time group
        ws.cell(row=1, column=1, value=tg.name)

        # Sort tickets
        tickets = sorted(
            tg.tickets, key=lambda ticket: ticket.lastname.lower())

        # Set ticket offset
        i = 0
        for t in tickets:
            # Filter cancelled tickets
            if t.status == TicketStatusEnum.cancelled:
                continue

            # Add row with ticket
            row_number = FIRST_ROW_TO_UPDATE + i

            # Update WorkSheet with
            ws.cell(row=row_number, column=1, value=(i+1))
            ws.cell(row=row_number, column=2, value=t.lastname)  # Lastname
            ws.cell(row=row_number, column=3, value=t.firstname)  # Firstname
            ws.cell(row=row_number, column=4, value=t.email)     # E-mail

            # Increase ticket offset
            i += 1

    # Remove default styled sheet
    if "default" in wb.sheetnames:
        del wb['default']

    # Return workbook
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        return BytesIO(tmp.read())
